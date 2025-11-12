from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods, require_safe, require_POST
from django.http import JsonResponse, HttpResponse
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Sum, Q
from decimal import Decimal
from empresas.middleware import EmpresaFilterMixin
from .models import ReporteGenerado, ConfiguracionReporte
from contabilidad.models import Asiento, CuentaContable, Partida
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from datetime import datetime

# Constantes para evitar duplicación de literales de URL
REPORTES_INDEX_URL = 'reportes:index'
CONFIGURACIONES_LIST_URL = 'reportes:configuraciones_lista'
HISTORIAL_URL = 'reportes:historial'

# Constantes para formatos Excel
EXCEL_MONEY_FORMAT = '"$"#,##0.00'
EXCEL_TOTALES_LABEL = 'TOTALES:'

# Vistas temporales básicas
class ReportesIndexView(LoginRequiredMixin, TemplateView):
    template_name = 'reportes/index.html'

class LibroDiarioView(LoginRequiredMixin, TemplateView):
    template_name = 'reportes/diario.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa_activa = getattr(self.request, 'empresa_activa', None)
        
        # Obtener parámetros de filtro
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')
        
        # Consultar asientos confirmados
        queryset = Asiento.objects.filter(estado='confirmado')
        if empresa_activa:
            queryset = queryset.filter(empresa=empresa_activa)
        
        if fecha_inicio:
            queryset = queryset.filter(fecha_asiento__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(fecha_asiento__lte=fecha_fin)
        
        context['asientos'] = queryset.select_related('creado_por', 'empresa').prefetch_related('partidas__cuenta').order_by('fecha_asiento', 'numero_asiento')
        context['fecha_inicio'] = fecha_inicio
        context['fecha_fin'] = fecha_fin
        
        return context

class LibroMayorView(LoginRequiredMixin, TemplateView):
    template_name = 'reportes/mayor.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa_activa = getattr(self.request, 'empresa_activa', None)
        
        # Obtener cuentas contables para selección
        cuentas = CuentaContable.objects.filter(acepta_movimiento=True, activa=True)
        if empresa_activa:
            cuentas = cuentas.filter(empresa=empresa_activa)
        
        context['cuentas'] = cuentas.order_by('codigo')
        
        return context

class LibroMayorCuentaView(LoginRequiredMixin, TemplateView):
    template_name = 'reportes/mayor_cuenta.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa_activa = getattr(self.request, 'empresa_activa', None)
        cuenta_pk = kwargs.get('cuenta_pk')
        
        # Obtener la cuenta
        try:
            cuenta = CuentaContable.objects.get(pk=cuenta_pk)
            if empresa_activa and cuenta.empresa != empresa_activa:
                messages.error(self.request, 'La cuenta no pertenece a la empresa activa.')
                return context
        except CuentaContable.DoesNotExist:
            messages.error(self.request, 'La cuenta no existe.')
            return context
        
        # Obtener parámetros de filtro
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')
        
        # Consultar partidas de la cuenta en asientos confirmados
        queryset = Partida.objects.filter(
            cuenta=cuenta,
            asiento__estado='confirmado'
        )
        
        if fecha_inicio:
            queryset = queryset.filter(asiento__fecha_asiento__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(asiento__fecha_asiento__lte=fecha_fin)
        
        partidas = queryset.select_related('asiento', 'cuenta').order_by('asiento__fecha_asiento', 'asiento__numero_asiento', 'orden')
        
        context['cuenta'] = cuenta
        context['partidas'] = partidas
        context['fecha_inicio'] = fecha_inicio
        context['fecha_fin'] = fecha_fin
        
        return context

class BalanceComprobacionView(LoginRequiredMixin, TemplateView):
    template_name = 'reportes/balance_comprobacion.html'
    
    def _calcular_agregados_cuenta(self, cuenta, fecha_corte):
        """Calcula débitos y créditos totales de una cuenta."""
        agregado = Partida.objects.filter(
            cuenta=cuenta,
            asiento__estado='confirmado',
            asiento__fecha_asiento__lte=fecha_corte
        ).aggregate(
            sum_debito=Sum('valor_debito'),
            sum_credito=Sum('valor_credito')
        )
        
        return (
            agregado['sum_debito'] or Decimal('0.00'),
            agregado['sum_credito'] or Decimal('0.00')
        )
    
    def _incluir_saldo_inicial(self, cuenta, total_debito, total_credito):
        """Agrega el saldo inicial según la naturaleza de la cuenta."""
        if cuenta.naturaleza == 'D':
            return total_debito + cuenta.saldo_inicial, total_credito
        return total_debito, total_credito + cuenta.saldo_inicial
    
    def _calcular_saldos_naturaleza(self, cuenta, total_debito, total_credito):
        """Calcula saldos deudor y acreedor según naturaleza."""
        if cuenta.naturaleza == 'D':
            saldo = total_debito - total_credito
            return (
                saldo if saldo > 0 else Decimal('0.00'),
                abs(saldo) if saldo < 0 else Decimal('0.00')
            )
        
        saldo = total_credito - total_debito
        return (
            abs(saldo) if saldo < 0 else Decimal('0.00'),
            saldo if saldo > 0 else Decimal('0.00')
        )
    
    def _procesar_cuenta_balance(self, cuenta, fecha_corte):
        """Procesa una cuenta para el balance de comprobación."""
        total_debito, total_credito = self._calcular_agregados_cuenta(cuenta, fecha_corte)
        total_debito, total_credito = self._incluir_saldo_inicial(
            cuenta, total_debito, total_credito
        )
        
        # Solo incluir cuentas con movimiento
        if total_debito == 0 and total_credito == 0:
            return None
        
        saldo_deudor, saldo_acreedor = self._calcular_saldos_naturaleza(
            cuenta, total_debito, total_credito
        )
        
        return {
            'cuenta': cuenta,
            'total_debito': total_debito,
            'total_credito': total_credito,
            'saldo_deudor': saldo_deudor,
            'saldo_acreedor': saldo_acreedor
        }
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa_activa = getattr(self.request, 'empresa_activa', None)
        
        fecha_corte = self.request.GET.get('fecha_corte')
        tipo_cuenta = self.request.GET.get('tipo_cuenta', '')
        
        if not fecha_corte:
            context.update({
                'fecha_corte': None,
                'tipo_cuenta': tipo_cuenta,
                'cuentas': []
            })
            return context
        
        # Obtener cuentas activas filtradas
        cuentas_query = CuentaContable.objects.filter(activa=True)
        if empresa_activa:
            cuentas_query = cuentas_query.filter(empresa=empresa_activa)
        if tipo_cuenta:
            cuentas_query = cuentas_query.filter(tipo_cuenta=tipo_cuenta)
        
        # Procesar cuentas
        cuentas_con_saldo = [
            cuenta_data
            for cuenta in cuentas_query.order_by('codigo')
            if (cuenta_data := self._procesar_cuenta_balance(cuenta, fecha_corte))
        ]
        
        context.update({
            'cuentas': cuentas_con_saldo,
            'fecha_corte': fecha_corte,
            'tipo_cuenta': tipo_cuenta
        })
        
        return context

class EstadoResultadosView(LoginRequiredMixin, TemplateView):
    template_name = 'reportes/estado_resultados.html'
    
    def _calcular_totales_cuenta(self, cuenta, fecha_inicio, fecha_fin):
        """Calcula débitos y créditos de una cuenta en un período."""
        agregado = Partida.objects.filter(
            cuenta=cuenta,
            asiento__estado='confirmado',
            asiento__fecha_asiento__gte=fecha_inicio,
            asiento__fecha_asiento__lte=fecha_fin
        ).aggregate(
            sum_debito=Sum('valor_debito'),
            sum_credito=Sum('valor_credito')
        )
        
        return (
            agregado['sum_debito'] or Decimal('0.00'),
            agregado['sum_credito'] or Decimal('0.00')
        )
    
    def _calcular_saldo_por_tipo(self, cuenta, total_debito, total_credito):
        """Calcula el saldo según el tipo de cuenta."""
        if cuenta.tipo_cuenta == 'INGRESO':
            return total_credito - total_debito
        # COSTO y GASTO usan la misma fórmula
        return total_debito - total_credito
    
    def _procesar_cuenta(self, cuenta, fecha_inicio, fecha_fin):
        """Procesa una cuenta y retorna su saldo si es positivo."""
        total_debito, total_credito = self._calcular_totales_cuenta(
            cuenta, fecha_inicio, fecha_fin
        )
        
        saldo = self._calcular_saldo_por_tipo(cuenta, total_debito, total_credito)
        
        if saldo > 0:
            cuenta.saldo = saldo
            return cuenta
        return None
    
    def _clasificar_cuentas(self, cuentas_query, fecha_inicio, fecha_fin):
        """Clasifica cuentas en ingresos, costos y gastos."""
        ingresos, costos, gastos = [], [], []
        
        for cuenta in cuentas_query:
            cuenta_procesada = self._procesar_cuenta(cuenta, fecha_inicio, fecha_fin)
            if cuenta_procesada:
                if cuenta.tipo_cuenta == 'INGRESO':
                    ingresos.append(cuenta_procesada)
                elif cuenta.tipo_cuenta == 'COSTO':
                    costos.append(cuenta_procesada)
                elif cuenta.tipo_cuenta == 'GASTO':
                    gastos.append(cuenta_procesada)
        
        return ingresos, costos, gastos
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa_activa = getattr(self.request, 'empresa_activa', None)
        
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            context.update({
                'fecha_inicio': None,
                'fecha_fin': None,
                'ingresos': [],
                'costos': [],
                'gastos': []
            })
            return context
        
        # Obtener cuentas activas
        cuentas_query = CuentaContable.objects.filter(activa=True)
        if empresa_activa:
            cuentas_query = cuentas_query.filter(empresa=empresa_activa)
        
        # Clasificar cuentas
        ingresos, costos, gastos = self._clasificar_cuentas(
            cuentas_query, fecha_inicio, fecha_fin
        )
        
        context.update({
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'ingresos': sorted(ingresos, key=lambda x: x.codigo),
            'costos': sorted(costos, key=lambda x: x.codigo),
            'gastos': sorted(gastos, key=lambda x: x.codigo)
        })
        
        return context

class BalanceGeneralView(LoginRequiredMixin, TemplateView):
    template_name = 'reportes/balance_general.html'
    
    def _calcular_saldo_cuenta(self, cuenta, fecha_corte):
        """Calcula el saldo de una cuenta hasta la fecha de corte."""
        agregado = Partida.objects.filter(
            cuenta=cuenta,
            asiento__estado='confirmado',
            asiento__fecha_asiento__lte=fecha_corte
        ).aggregate(
            sum_debito=Sum('valor_debito'),
            sum_credito=Sum('valor_credito')
        )
        
        total_debito = agregado['sum_debito'] or Decimal('0.00')
        total_credito = agregado['sum_credito'] or Decimal('0.00')
        
        # Incluir saldo inicial
        if cuenta.naturaleza == 'D':
            total_debito += cuenta.saldo_inicial
        else:
            total_credito += cuenta.saldo_inicial
        
        # Calcular saldo según naturaleza
        if cuenta.naturaleza == 'D':
            return total_debito - total_credito
        return total_credito - total_debito
    
    def _obtener_prefijo_codigo(self, cuenta):
        """Obtiene los primeros 2 dígitos del código de cuenta."""
        if len(cuenta.codigo) >= 2 and cuenta.codigo[:2].isdigit():
            return int(cuenta.codigo[:2])
        return 0
    
    def _clasificar_activo(self, cuenta, activos_corrientes, activos_no_corrientes):
        """Clasifica una cuenta de activo en corriente o no corriente."""
        codigo_num = self._obtener_prefijo_codigo(cuenta)
        if codigo_num <= 13:
            activos_corrientes.append(cuenta)
        else:
            activos_no_corrientes.append(cuenta)
    
    def _clasificar_pasivo(self, cuenta, pasivos_corrientes, pasivos_no_corrientes):
        """Clasifica una cuenta de pasivo en corriente o no corriente."""
        codigo_num = self._obtener_prefijo_codigo(cuenta)
        if codigo_num <= 23:
            pasivos_corrientes.append(cuenta)
        else:
            pasivos_no_corrientes.append(cuenta)
    
    def _clasificar_cuenta_balance(self, cuenta, clasificaciones):
        """Clasifica una cuenta en la categoría correspondiente."""
        if cuenta.tipo_cuenta == 'ACTIVO':
            self._clasificar_activo(
                cuenta,
                clasificaciones['activos_corrientes'],
                clasificaciones['activos_no_corrientes']
            )
        elif cuenta.tipo_cuenta == 'PASIVO':
            self._clasificar_pasivo(
                cuenta,
                clasificaciones['pasivos_corrientes'],
                clasificaciones['pasivos_no_corrientes']
            )
        elif cuenta.tipo_cuenta == 'PATRIMONIO':
            clasificaciones['patrimonio'].append(cuenta)
    
    def _procesar_cuentas_balance(self, cuentas_query, fecha_corte):
        """Procesa y clasifica todas las cuentas para el balance."""
        clasificaciones = {
            'activos_corrientes': [],
            'activos_no_corrientes': [],
            'pasivos_corrientes': [],
            'pasivos_no_corrientes': [],
            'patrimonio': []
        }
        
        for cuenta in cuentas_query:
            saldo = self._calcular_saldo_cuenta(cuenta, fecha_corte)
            
            if saldo != 0:
                cuenta.saldo = abs(saldo)
                self._clasificar_cuenta_balance(cuenta, clasificaciones)
        
        return clasificaciones
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa_activa = getattr(self.request, 'empresa_activa', None)
        
        fecha_corte = self.request.GET.get('fecha_corte')
        
        if not fecha_corte:
            context.update({
                'fecha_corte': None,
                'activos_corrientes': [],
                'activos_no_corrientes': [],
                'pasivos_corrientes': [],
                'pasivos_no_corrientes': [],
                'patrimonio': []
            })
            return context
        
        # Obtener cuentas activas
        cuentas_query = CuentaContable.objects.filter(activa=True)
        if empresa_activa:
            cuentas_query = cuentas_query.filter(empresa=empresa_activa)
        
        # Procesar y clasificar cuentas
        clasificaciones = self._procesar_cuentas_balance(cuentas_query, fecha_corte)
        
        context.update({
            'fecha_corte': fecha_corte,
            'activos_corrientes': sorted(clasificaciones['activos_corrientes'], key=lambda x: x.codigo),
            'activos_no_corrientes': sorted(clasificaciones['activos_no_corrientes'], key=lambda x: x.codigo),
            'pasivos_corrientes': sorted(clasificaciones['pasivos_corrientes'], key=lambda x: x.codigo),
            'pasivos_no_corrientes': sorted(clasificaciones['pasivos_no_corrientes'], key=lambda x: x.codigo),
            'patrimonio': sorted(clasificaciones['patrimonio'], key=lambda x: x.codigo)
        })
        
        return context

class FlujoEfectivoView(LoginRequiredMixin, TemplateView):
    template_name = 'reportes/flujo_efectivo.html'

class ConfiguracionReporteListView(LoginRequiredMixin, EmpresaFilterMixin, ListView):
    model = ConfiguracionReporte
    template_name = 'reportes/configuraciones_lista.html'
    context_object_name = 'object_list'
    paginate_by = 20
    
    def get_queryset(self):
        return super().get_queryset().select_related('creado_por', 'empresa').order_by('tipo_reporte', 'nombre')

class ConfiguracionReporteDetailView(LoginRequiredMixin, EmpresaFilterMixin, DetailView):
    model = ConfiguracionReporte
    template_name = 'reportes/configuraciones_detalle.html'

class ConfiguracionReporteCreateView(LoginRequiredMixin, EmpresaFilterMixin, CreateView):
    model = ConfiguracionReporte
    template_name = 'reportes/configuraciones_crear.html'
    fields = ['nombre', 'tipo_reporte', 'descripcion', 'configuracion', 'es_publica']
    success_url = reverse_lazy(CONFIGURACIONES_LIST_URL)
    
    def form_valid(self, form):
        form.instance.empresa = getattr(self.request, 'empresa_activa', None)
        form.instance.creado_por = self.request.user
        messages.success(self.request, f'Configuración "{form.instance.nombre}" creada exitosamente.')
        return super().form_valid(form)

class ConfiguracionReporteUpdateView(LoginRequiredMixin, EmpresaFilterMixin, UpdateView):
    model = ConfiguracionReporte
    template_name = 'reportes/configuraciones_editar.html'
    fields = ['nombre', 'tipo_reporte', 'descripcion', 'configuracion', 'es_publica']
    success_url = reverse_lazy(CONFIGURACIONES_LIST_URL)
    
    def form_valid(self, form):
        messages.success(self.request, f'Configuración "{form.instance.nombre}" actualizada exitosamente.')
        return super().form_valid(form)

class ConfiguracionReporteDeleteView(LoginRequiredMixin, EmpresaFilterMixin, DeleteView):
    model = ConfiguracionReporte
    template_name = 'reportes/configuraciones_eliminar.html'
    success_url = reverse_lazy(CONFIGURACIONES_LIST_URL)
    
    def delete(self, request, *args, **kwargs):
        config = self.get_object()
        messages.success(request, f'Configuración "{config.nombre}" eliminada exitosamente.')
        return super().delete(request, *args, **kwargs)

class ReporteGeneradoListView(LoginRequiredMixin, EmpresaFilterMixin, ListView):
    model = ReporteGenerado
    template_name = 'reportes/historial.html'
    context_object_name = 'object_list'
    paginate_by = 30
    
    def get_queryset(self):
        return super().get_queryset().select_related('generado_por', 'empresa').order_by('-fecha_generacion')

class ReporteGeneradoDetailView(LoginRequiredMixin, EmpresaFilterMixin, DetailView):
    model = ReporteGenerado
    template_name = 'reportes/historial_detalle.html'

@login_required
@require_safe
def generar_libro_diario(request):
    return redirect('reportes:diario')

@login_required
@require_http_methods(["GET"])
def exportar_libro_diario(request):
    """Exportar libro diario a Excel o PDF"""
    empresa_activa = getattr(request, 'empresa_activa', None)
    formato = request.GET.get('formato', 'excel')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if not fecha_inicio or not fecha_fin:
        return HttpResponse("Fechas de inicio y fin requeridas", status=400)
    
    # Obtener asientos del periodo
    asientos = Asiento.objects.filter(
        empresa=empresa_activa,
        fecha_asiento__gte=fecha_inicio,
        fecha_asiento__lte=fecha_fin
    ).select_related('creado_por').prefetch_related('partidas__cuenta').order_by('fecha_asiento', 'numero_asiento')
    
    if formato == 'excel':
        return _exportar_diario_excel(asientos, fecha_inicio, fecha_fin, empresa_activa)
    elif formato == 'pdf':
        return _exportar_diario_pdf(asientos, fecha_inicio, fecha_fin, empresa_activa)
    else:
        return HttpResponse("Formato no soportado", status=400)

def _exportar_diario_excel(asientos, fecha_inicio, fecha_fin, empresa):
    """Generar archivo Excel del libro diario"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Diario"
    
    # Estilos
    titulo_font = Font(size=14, bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = f'{empresa.razon_social} - Libro Diario'
    ws['A1'].font = titulo_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f'Periodo: {fecha_inicio} a {fecha_fin}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Encabezados
    headers = ['Fecha', 'Asiento', 'Cuenta', 'Concepto', 'Débito', 'Crédito', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Datos
    row = 5
    total_debitos = Decimal('0.00')
    total_creditos = Decimal('0.00')
    
    for asiento in asientos:
        # Partidas del asiento
        for partida in asiento.partidas.all():
            ws.cell(row=row, column=1, value=asiento.fecha_asiento.strftime('%Y-%m-%d')).border = border
            ws.cell(row=row, column=2, value=asiento.numero_asiento).border = border
            ws.cell(row=row, column=3, value=f"{partida.cuenta.codigo} - {partida.cuenta.nombre}").border = border
            ws.cell(row=row, column=4, value=partida.concepto or asiento.concepto).border = border
            ws.cell(row=row, column=5, value=float(partida.valor_debito)).border = border
            ws.cell(row=row, column=6, value=float(partida.valor_credito)).border = border
            ws.cell(row=row, column=7, value=asiento.get_estado_display()).border = border
            
            # Formato de moneda
            ws.cell(row=row, column=5).number_format = EXCEL_MONEY_FORMAT
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=6).number_format = EXCEL_MONEY_FORMAT
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
            
            total_debitos += partida.valor_debito
            total_creditos += partida.valor_credito
            
            row += 1
    
    # Totales
    ws.cell(row=row, column=1, value=EXCEL_TOTALES_LABEL).font = Font(bold=True)
    ws.cell(row=row, column=5, value=float(total_debitos)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(total_creditos)).font = Font(bold=True)
    
    ws.cell(row=row, column=5).number_format = EXCEL_MONEY_FORMAT
    ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=5).border = border
    ws.cell(row=row, column=6).number_format = EXCEL_MONEY_FORMAT
    ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=6).border = border
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=libro_diario_{fecha_inicio}_{fecha_fin}.xlsx'
    return response

def _exportar_diario_pdf(asientos, fecha_inicio, fecha_fin, empresa):
    """Generar archivo PDF del libro diario"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(f'{empresa.razon_social}', titulo_style))
    elements.append(Paragraph('Libro Diario', titulo_style))
    elements.append(Paragraph(f'Periodo: {fecha_inicio} a {fecha_fin}', styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabla de datos
    data = [['Fecha', 'Asiento', 'Cuenta', 'Concepto', 'Débito', 'Crédito', 'Estado']]
    
    total_debitos = Decimal('0.00')
    total_creditos = Decimal('0.00')
    
    for asiento in asientos:
        for partida in asiento.partidas.all():
            data.append([
                asiento.fecha_asiento.strftime('%Y-%m-%d'),
                asiento.numero_asiento,
                f"{partida.cuenta.codigo}\n{partida.cuenta.nombre[:25]}",
                (partida.concepto or asiento.concepto)[:30],
                f"${partida.valor_debito:,.2f}",
                f"${partida.valor_credito:,.2f}",
                asiento.get_estado_display()[:8]
            ])
            total_debitos += partida.valor_debito
            total_creditos += partida.valor_credito
    
    # Fila de totales
    data.append([
        EXCEL_TOTALES_LABEL,
        '',
        '',
        '',
        f"${total_debitos:,.2f}",
        f"${total_creditos:,.2f}",
        ''
    ])
    
    table = Table(data, colWidths=[0.8*inch, 1*inch, 2*inch, 2*inch, 1.2*inch, 1.2*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (3, -1), 'LEFT'),
        ('ALIGN', (4, 1), (5, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=libro_diario_{fecha_inicio}_{fecha_fin}.pdf'
    return response

@login_required
@require_safe
def generar_libro_mayor(request):
    return redirect('reportes:mayor')

@login_required
@require_http_methods(["GET"])
def exportar_libro_mayor(request):
    return HttpResponse("CSV")

@login_required
@require_safe
def generar_balance_comprobacion(request):
    return redirect('reportes:balance_comprobacion')

@login_required
@require_http_methods(["GET"])
def exportar_balance_comprobacion(request):
    """Exportar balance de comprobación a Excel o PDF"""
    # Validar parámetros de entrada
    fecha_corte = request.GET.get('fecha_corte')
    if not fecha_corte:
        return HttpResponse("Fecha de corte requerida", status=400)
    
    # Obtener datos del balance
    empresa_activa = getattr(request, 'empresa_activa', None)
    tipo_cuenta = request.GET.get('tipo_cuenta', '')
    formato = request.GET.get('formato', 'excel')
    
    balance_data = _obtener_datos_balance_comprobacion(empresa_activa, fecha_corte, tipo_cuenta)
    
    # Exportar según formato solicitado
    return _exportar_segun_formato(formato, balance_data, fecha_corte, empresa_activa)

def _obtener_datos_balance_comprobacion(empresa_activa, fecha_corte, tipo_cuenta):
    """Obtener y procesar datos del balance de comprobación"""
    cuentas = _obtener_cuentas_filtradas(empresa_activa, tipo_cuenta)
    
    cuentas_con_saldo = []
    totales_globales = _inicializar_totales_globales()
    
    for cuenta in cuentas.order_by('codigo'):
        datos_cuenta = _procesar_cuenta_individual(cuenta, fecha_corte)
        
        if _cuenta_tiene_movimientos(datos_cuenta):
            cuentas_con_saldo.append(datos_cuenta)
            _actualizar_totales_globales(totales_globales, datos_cuenta)
    
    return {
        'cuentas': cuentas_con_saldo,
        'totales': totales_globales
    }

def _obtener_cuentas_filtradas(empresa_activa, tipo_cuenta):
    """Obtener cuentas filtradas según criterios"""
    cuentas = CuentaContable.objects.filter(
        empresa=empresa_activa,
        activa=True,
        acepta_movimiento=True
    )
    
    if tipo_cuenta:
        cuentas = cuentas.filter(tipo_cuenta=tipo_cuenta)
    
    return cuentas

def _inicializar_totales_globales():
    """Inicializar estructura de totales globales"""
    return {
        'debitos': Decimal('0.00'),
        'creditos': Decimal('0.00'),
        'saldo_deudor': Decimal('0.00'),
        'saldo_acreedor': Decimal('0.00')
    }

def _procesar_cuenta_individual(cuenta, fecha_corte):
    """Procesar una cuenta individual y calcular sus saldos"""
    # Obtener movimientos de la cuenta
    agregado = Partida.objects.filter(
        cuenta=cuenta,
        asiento__estado='confirmado',
        asiento__fecha_asiento__lte=fecha_corte
    ).aggregate(
        sum_debito=Sum('valor_debito'),
        sum_credito=Sum('valor_credito')
    )
    
    total_debito = agregado['sum_debito'] or Decimal('0.00')
    total_credito = agregado['sum_credito'] or Decimal('0.00')
    
    # Incluir saldo inicial
    total_debito, total_credito = _incluir_saldo_inicial(cuenta, total_debito, total_credito)
    
    # Calcular saldos finales
    saldo_deudor, saldo_acreedor = _calcular_saldos_finales(cuenta, total_debito, total_credito)
    
    return {
        'codigo': cuenta.codigo,
        'nombre': cuenta.nombre,
        'tipo': cuenta.get_tipo_cuenta_display(),
        'debito': total_debito,
        'credito': total_credito,
        'saldo_deudor': saldo_deudor,
        'saldo_acreedor': saldo_acreedor
    }

def _incluir_saldo_inicial(cuenta, total_debito, total_credito):
    """Incluir saldo inicial según naturaleza de la cuenta"""
    if cuenta.naturaleza == 'debito':
        total_debito += cuenta.saldo_inicial
    else:
        total_credito += cuenta.saldo_inicial
    
    return total_debito, total_credito

def _calcular_saldos_finales(cuenta, total_debito, total_credito):
    """Calcular saldos deudor y acreedor según naturaleza de la cuenta"""
    if cuenta.naturaleza == 'debito':
        saldo = total_debito - total_credito
        saldo_deudor = saldo if saldo > 0 else Decimal('0.00')
        saldo_acreedor = abs(saldo) if saldo < 0 else Decimal('0.00')
    else:
        saldo = total_credito - total_debito
        saldo_acreedor = saldo if saldo > 0 else Decimal('0.00')
        saldo_deudor = abs(saldo) if saldo < 0 else Decimal('0.00')
    
    return saldo_deudor, saldo_acreedor

def _cuenta_tiene_movimientos(datos_cuenta):
    """Verificar si la cuenta tiene movimientos"""
    return datos_cuenta['debito'] > 0 or datos_cuenta['credito'] > 0

def _actualizar_totales_globales(totales_globales, datos_cuenta):
    """Actualizar totales globales con datos de una cuenta"""
    totales_globales['debitos'] += datos_cuenta['debito']
    totales_globales['creditos'] += datos_cuenta['credito']
    totales_globales['saldo_deudor'] += datos_cuenta['saldo_deudor']
    totales_globales['saldo_acreedor'] += datos_cuenta['saldo_acreedor']

def _exportar_segun_formato(formato, balance_data, fecha_corte, empresa_activa):
    """Exportar balance según formato solicitado"""
    cuentas = balance_data['cuentas']
    totales = balance_data['totales']
    
    if formato == 'excel':
        return _exportar_balance_excel(
            cuentas, fecha_corte, empresa_activa,
            totales['debitos'], totales['creditos'],
            totales['saldo_deudor'], totales['saldo_acreedor']
        )
    elif formato == 'pdf':
        return _exportar_balance_pdf(
            cuentas, fecha_corte, empresa_activa,
            totales['debitos'], totales['creditos'],
            totales['saldo_deudor'], totales['saldo_acreedor']
        )
    else:
        return HttpResponse("Formato no soportado", status=400)

def _exportar_balance_excel(cuentas, fecha_corte, empresa, total_deb, total_cred, total_sd, total_sa):
    """Generar archivo Excel del balance de comprobación"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance de Comprobación"
    
    # Estilos
    titulo_font = Font(size=14, bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = f'{empresa.razon_social} - Balance de Comprobación'
    ws['A1'].font = titulo_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f'Fecha de Corte: {fecha_corte}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Encabezados
    headers = ['Código', 'Cuenta', 'Tipo', 'Débitos', 'Créditos', 'Saldo Deudor', 'Saldo Acreedor']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Datos
    row = 5
    for cuenta in cuentas:
        ws.cell(row=row, column=1, value=cuenta['codigo']).border = border
        ws.cell(row=row, column=2, value=cuenta['nombre']).border = border
        ws.cell(row=row, column=3, value=cuenta['tipo']).border = border
        ws.cell(row=row, column=4, value=float(cuenta['debito'])).border = border
        ws.cell(row=row, column=5, value=float(cuenta['credito'])).border = border
        ws.cell(row=row, column=6, value=float(cuenta['saldo_deudor'])).border = border
        ws.cell(row=row, column=7, value=float(cuenta['saldo_acreedor'])).border = border
        
        # Formato de moneda
        for col in range(4, 8):
            ws.cell(row=row, column=col).number_format = EXCEL_MONEY_FORMAT
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        
        row += 1
    
    # Totales
    ws.cell(row=row, column=1, value=EXCEL_TOTALES_LABEL).font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(total_deb)).font = Font(bold=True)
    ws.cell(row=row, column=5, value=float(total_cred)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(total_sd)).font = Font(bold=True)
    ws.cell(row=row, column=7, value=float(total_sa)).font = Font(bold=True)
    
    for col in range(4, 8):
        ws.cell(row=row, column=col).number_format = EXCEL_MONEY_FORMAT
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=col).border = border
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=balance_comprobacion_{fecha_corte}.xlsx'
    return response

def _exportar_balance_pdf(cuentas, fecha_corte, empresa, total_deb, total_cred, total_sd, total_sa):
    """Generar archivo PDF del balance de comprobación"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(f'{empresa.razon_social}', titulo_style))
    elements.append(Paragraph('Balance de Comprobación', titulo_style))
    elements.append(Paragraph(f'Fecha de Corte: {fecha_corte}', styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabla de datos
    data = [['Código', 'Cuenta', 'Tipo', 'Débitos', 'Créditos', 'Saldo Deudor', 'Saldo Acreedor']]
    
    for cuenta in cuentas:
        data.append([
            cuenta['codigo'],
            cuenta['nombre'][:35],  # Truncar nombres largos
            cuenta['tipo'][:10],
            f"${cuenta['debito']:,.2f}",
            f"${cuenta['credito']:,.2f}",
            f"${cuenta['saldo_deudor']:,.2f}",
            f"${cuenta['saldo_acreedor']:,.2f}"
        ])
    
    # Fila de totales
    data.append([
        EXCEL_TOTALES_LABEL,
        '',
        '',
        f"${total_deb:,.2f}",
        f"${total_cred:,.2f}",
        f"${total_sd:,.2f}",
        f"${total_sa:,.2f}"
    ])
    
    table = Table(data, colWidths=[0.8*inch, 2.5*inch, 0.8*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=balance_comprobacion_{fecha_corte}.pdf'
    return response

@login_required
@require_safe
def generar_estado_resultados(request):
    return redirect('reportes:estado_resultados')

@login_required
@require_http_methods(["GET"])
def exportar_estado_resultados(request):
    return HttpResponse("CSV")

@login_required
@require_safe
def generar_balance_general(request):
    return redirect('reportes:balance_general')

@login_required
@require_http_methods(["GET"])
def exportar_balance_general(request):
    return HttpResponse("CSV")

@login_required
@require_safe
def generar_flujo_efectivo(request):
    return redirect('reportes:flujo_efectivo')

@login_required
@require_http_methods(["GET"])
def exportar_flujo_efectivo(request):
    return HttpResponse("CSV")

@login_required
@require_http_methods(["POST"])
def usar_configuracion(request, pk):
    return redirect('reportes:index')

@login_required
@require_http_methods(["GET"])
def descargar_reporte(request, pk):
    return HttpResponse("Archivo")

@login_required
@require_http_methods(["GET"])
def validar_periodo_reporte(request):
    return JsonResponse({'valido': True})

@login_required
@require_safe
def preview_reporte(request):
    return JsonResponse({'preview': {}})
